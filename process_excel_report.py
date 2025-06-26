import openpyxl
import os

ORIGINAL_EXCEL_FILE_NAME = 'mortgage_rates_report.xlsx'
EMAILED_EXCEL_FILE_NAME = 'mortgage_rates_report_emailed.xlsx'
ORIGINAL_DATA_SHEET_NAME = 'Mortgage Rates' # As created by export_to_excel.py
NEW_AGGREGATE_SHEET_NAME = 'Aggregate Report'

def create_emailed_report_with_average():
    """
    Loads the original Excel report, adds a new sheet with the average rate,
    and saves it as a new file.
    """
    if not os.path.exists(ORIGINAL_EXCEL_FILE_NAME):
        print(f"Error: Original Excel file '{ORIGINAL_EXCEL_FILE_NAME}' not found.")
        print("Please run export_to_excel.py first to generate it.")
        return

    try:
        # Load the existing workbook
        workbook = openpyxl.load_workbook(ORIGINAL_EXCEL_FILE_NAME)
    except Exception as e:
        print(f"Error loading Excel file '{ORIGINAL_EXCEL_FILE_NAME}': {e}")
        return

    # Get the original data sheet
    if ORIGINAL_DATA_SHEET_NAME not in workbook.sheetnames:
        print(f"Error: Sheet '{ORIGINAL_DATA_SHEET_NAME}' not found in the workbook.")
        return
    data_sheet = workbook[ORIGINAL_DATA_SHEET_NAME]

    # Determine the range for the average calculation in the "Rate" column (Column B)
    # Assuming headers are in row 1, data starts from row 2.
    # data_sheet.max_row gives the last row index that contains data.
    if data_sheet.max_row < 2:
        print(f"No data rows found in sheet '{ORIGINAL_DATA_SHEET_NAME}' to calculate average.")
        # Still create the aggregate sheet, but the average will likely be an error or zero.
        last_data_row = 1 # To avoid issues with range B2:B1
    else:
        last_data_row = data_sheet.max_row
    
    rate_column_letter = 'B' # 'Rate' is in column B as per export_to_excel.py
    average_formula_range = f"'{ORIGINAL_DATA_SHEET_NAME}'!{rate_column_letter}2:{rate_column_letter}{last_data_row}"
    
    # Create the new "Aggregate Report" sheet
    if NEW_AGGREGATE_SHEET_NAME in workbook.sheetnames:
        # If it somehow exists, remove it to start fresh or use it
        print(f"Sheet '{NEW_AGGREGATE_SHEET_NAME}' already exists. Recreating it.")
        del workbook[NEW_AGGREGATE_SHEET_NAME]
    
    aggregate_sheet = workbook.create_sheet(NEW_AGGREGATE_SHEET_NAME)

    # Add headers/labels to the aggregate sheet
    aggregate_sheet['A1'] = "Metric"
    aggregate_sheet['B1'] = "Value"
    
    aggregate_sheet['A2'] = "Average Original Rate"
    
    # Add the Excel formula to calculate the average
    # The formula will be =AVERAGE('Mortgage Rates'!B2:B<last_row_number>)
    aggregate_sheet['B2'] = f"=AVERAGE({average_formula_range})"
    aggregate_sheet['B2'].number_format = '0.00' # Format the result cell

    # Adjust column widths for the aggregate sheet
    aggregate_sheet.column_dimensions['A'].width = 25
    aggregate_sheet.column_dimensions['B'].width = 15

    try:
        # Save the workbook to the new file name
        workbook.save(EMAILED_EXCEL_FILE_NAME)
        print(f"Successfully created '{EMAILED_EXCEL_FILE_NAME}' with the aggregate report sheet.")
    except Exception as e:
        print(f"Error saving Excel file '{EMAILED_EXCEL_FILE_NAME}': {e}")

if __name__ == '__main__':
    create_emailed_report_with_average()
    print("Excel processing for email report complete.")

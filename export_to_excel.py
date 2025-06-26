import duckdb
import openpyxl
from openpyxl.utils import get_column_letter
import os
# No need to import 'date' from datetime specifically if only using for type hints
# or if DuckDB returns datetime.date objects which openpyxl handles.

DUCKDB_DATABASE_NAME = "homes_olap.duckdb"
MORTGAGE_RATES_TABLE_NAME = "mortgage_rates"
EXCEL_FILE_NAME = "mortgage_rates_report.xlsx"


def read_data_from_duckdb():
    """Reads all data from the mortgage_rates table in DuckDB."""
    if not os.path.exists(DUCKDB_DATABASE_NAME):
        print(f"Error: DuckDB database '{DUCKDB_DATABASE_NAME}' not found.")
        return None

    con = duckdb.connect(database=DUCKDB_DATABASE_NAME, read_only=True)
    print(f"Reading data from DuckDB table '{MORTGAGE_RATES_TABLE_NAME}'...")
    # Ensure we select date and rate
    query = f"SELECT date, rate FROM {MORTGAGE_RATES_TABLE_NAME} ORDER BY date"
    try:
        # DuckDB's fetchall() returns a list of tuples.
        # Each tuple contains values in the order of selection, e.g., (date_obj, rate_float)
        data = con.execute(query).fetchall()
        print(f"Successfully read {len(data)} rows from DuckDB.")
    except Exception as e:
        print(f"Error reading from DuckDB: {e}")
        data = None
    finally:
        con.close()

    return data


def write_to_excel(data):
    """Writes data to an Excel file with a calculated 'adjusted_rate' column."""
    if not data:
        print("No data provided to write to Excel.")
        return

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Mortgage Rates"

    # Write headers
    headers = ["Date", "Rate", "Adjusted Rate"]
    sheet.append(headers)

    # Write data rows and formula
    for row_idx, db_row_data in enumerate(
        data, start=2
    ):  # start=2 because Excel rows are 1-indexed and row 1 is header
        # db_row_data is a tuple, e.g., (datetime.date(2025, 6, 1), 6.85)
        # The date from DuckDB should be a datetime.date object, which openpyxl handles.
        date_value = db_row_data[0]
        rate_value = db_row_data[1]

        # Write date (column A) and rate (column B)
        sheet[f"A{row_idx}"] = date_value
        sheet[f"B{row_idx}"] = rate_value

        # Create formula for Adjusted Rate (column C)
        # Excel's WEEKDAY function: WEEKDAY(serial_number, [return_type])
        # Default return_type: 1 (Sunday) through 7 (Saturday).
        # So, weekend days are 1 (Sunday) or 7 (Saturday).
        date_cell_ref = f"A{row_idx}"
        rate_cell_ref = f"B{row_idx}"
        formula = f"=IF(OR(WEEKDAY({date_cell_ref})=1, WEEKDAY({date_cell_ref})=7), {rate_cell_ref}/2, {rate_cell_ref})"
        sheet[f"C{row_idx}"] = formula

    # Adjust column widths and number formats for better readability
    for col_idx, column_title in enumerate(headers, start=1):
        column_letter = get_column_letter(col_idx)
        if column_title == "Date":
            sheet.column_dimensions[column_letter].width = 12
            # Dates are typically formatted by Excel automatically, but can be set if needed
            # for row in range(2, sheet.max_row + 1):
            #     sheet[f"{column_letter}{row}"].number_format = 'YYYY-MM-DD' # Or your preferred date format
        elif column_title == "Rate":
            sheet.column_dimensions[column_letter].width = 10
            for row in range(2, sheet.max_row + 1):
                sheet[f"{column_letter}{row}"].number_format = "0.00"
        elif column_title == "Adjusted Rate":
            sheet.column_dimensions[column_letter].width = 15
            for row in range(2, sheet.max_row + 1):
                sheet[f"{column_letter}{row}"].number_format = "0.00"

    try:
        workbook.save(EXCEL_FILE_NAME)
        print(f"Data successfully written to '{EXCEL_FILE_NAME}'.")
        print(
            "The 'Adjusted Rate' column contains an Excel formula to calculate rates based on the day of the week."
        )
    except Exception as e:
        print(f"Error saving Excel file: {e}")


if __name__ == "__main__":
    # This script assumes cdc_to_duckdb.py has been run successfully and homes_olap.duckdb exists.

    duckdb_data = read_data_from_duckdb()
    if duckdb_data:
        write_to_excel(duckdb_data)
        print("Export to Excel process complete.")
    else:
        print(
            "Export to Excel process aborted due to missing DuckDB data or read error."
        )
